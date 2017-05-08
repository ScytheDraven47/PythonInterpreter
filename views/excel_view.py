# Ben Rogers-McKee (ScytheDraven47)
# 08/03/2017
#
# ExcelView for Interpreter Program

from openpyxl import Workbook
from openpyxl import load_workbook

from views.view import IView


class ExcelView(IView):

    def __init__(self):
        self.has_loaded_excel = False
        self.has_created_excel = False
        self.has_received_data_from_excel = False
        self.has_filled_excel = False
        self.has_saved_excel = False

    def get_data(self, excel_file):
        """Gets a single row of data from the excel file, returns the data as an array"""
        wb = load_workbook(filename=excel_file)
        self.has_loaded_excel = True
        ws = wb['input']
        starting_col = 1
        all_data = []
        for r in range(2, ws.max_row):
            row_data = []
            for c in range(starting_col, ws.max_column + 1):
                row_data.append(ws.cell(row=r, column=c).value)
            all_data.append(row_data)
        self.has_received_data_from_excel = True
        return all_data

    def output(self, message, excel_file):
        """Loads the excel file, saves data(message) into the first empty row, saves file"""
        try:
            wb = load_workbook(excel_file)
            ws = wb['output']
            self.has_loaded_excel = True
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "output"
            self.has_created_excel = True
        row = 1
        while ws['A'+str(row)].value is not None:
            row += 1
        col = 1
        for row_data in message:
            for value in row_data.values():
                ws.cell(row=row, column=col).value = value
                col += 1
            row += 1
            self.has_filled_excel = True
        wb.save(excel_file)
        self.has_saved_excel = True
